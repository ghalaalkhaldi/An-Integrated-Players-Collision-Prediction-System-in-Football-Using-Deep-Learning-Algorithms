from flask import Flask, render_template, request, redirect, url_for, send_from_directory, send_file
from flask_cors import CORS, cross_origin
from flask_socketio import SocketIO, emit
from werkzeug.utils import secure_filename
from flask_uploads import UploadSet
import os
import cv2
import pandas as pd
from datetime import datetime
from ultralytics import YOLO
import openpyxl
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
from openpyxl.styles import Alignment
from tracker import Tracker
import shutil

# Initialize the Flask application. `__name__` ensures the correct module is used.
app = Flask(__name__)
## Enable Cross-Origin Resource Sharing (CORS) to allow requests from different origins (e.g., frontend running on a different port)
cors = CORS(app)
#Integrate SocketIO into the Flask app to enable real-time communication
socketio = SocketIO(app)


#Define the directory to store uploaded media files (create it if it doesn't exist)
UPLOAD_FOLDER = 'static/uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
# Specify allowed file extensions for uploaded video
ALLOWED_EXTENSIONS = {'mp4', 'avi', 'mov', 'mkv'}

# Create a Flask-Uploads UploadSet named 'media' to handle media file uploads and set allowed extensions
media = UploadSet('media', ('mp4', 'avi', 'mov', 'mkv'))


#################################### HELPER FUNCTIONS #############################


# Check if there's a '.' in the filename (indicating an extension)
    # and if the last part after the last '.' (the extension itself) is in the allowed list.
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS



######################### ROUTING FUNCTIONS #########################

# Define the main route ('/') that handles both GET and POST requests
@app.route('/',methods=['GET','POST'])

# This function renders the main 'index.html' template, likely serving as
    #the initial landing page of the application
def main():
    return render_template('index.html') # Render the HTML template

# Define the '/submit' route to handle form submissions (GET and POST)
@app.route('/submit', methods=['GET','POST'])

# This function handles form submissions from the main page. It processes
    #uploaded video files, extracts form data (distance and confidence), and saves
  #  the video to the server.
def submit():

    # with open('./static/temp/file.txt', 'w') as file:
    #     file.write('working')


    if 'file' not in request.files:  # Check if a file was uploaded in the request
        print('No file part') # Log error if no file is found
        return redirect(request.url) # Redirect back to the submission page
    
    file = request.files['file'] # Get the uploaded file object
    
    # Extract distance and confidence values from the form data
    distance = int(request.form['distance'])
    confidence = str(request.form['confidence'])
    
    # (Debugging output to print values)
    print(distance, confidence)
    
    # Check if the file has a valid filename
    if file.filename == '':
        print('No image selected for uploading')
        return redirect(request.url)
    else:
    # Secure the filename to prevent malicious uploads
        filename = secure_filename(file.filename)
    # Create the full path to save the file in the uploads folder
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        print(filepath) # (Debugging output to print the save path)
        file.save(filepath)

        # Define styles
        bold_font = Font(bold=True)
        safe_font = Font(color='00FF00')  # Green
        collision_font = Font(color='FF0000')  # Red
        close_proximity_font = Font(color='FFFF00')  # Yellow
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow background

        # Input minimum distance threshold and confidence
        min_distance_threshold = int(distance)
        confidence_threshold = float(confidence)

        # Model setup
        model = YOLO('./static/model/best.pt')
        class_list = ["Human"]
        tracker = Tracker()

        vid_file_name = filepath
        ids_list = []
        ids_list_check = []
        cap = cv2.VideoCapture(vid_file_name)
        count = 0
        output_file = './static/output/output.mp4'
        frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        fps = cap.get(cv2.CAP_PROP_FPS)
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter(output_file, fourcc, fps, (frame_width, frame_height))

        overlap_threshold = 0.3  # Set your overlap threshold here in percent
        overlap_threshold_limit = 0.9
        warning_shown = False
        alert_shown = False

        log_data = []

        # Process video
        while True:
            ret, frame = cap.read()
            if not ret:
                break
            count += 1

            results = model.predict(frame, conf=confidence_threshold)
            a = results[0].boxes.cpu().data.numpy()
            px = pd.DataFrame(a).astype("float")
            list = []

            for index, row in px.iterrows():
                x1 = int(row[0])
                y1 = int(row[1])
                x2 = int(row[2])
                y2 = int(row[3])
                d = int(row[5])
                c = class_list[d]
                if "Human" in c:
                    list.append([x1, y1, x2, y2])

            bbox_id = tracker.update(list)
            
            warning_shown = False
            alert_shown = False
            
            for i, bbox1 in enumerate(bbox_id):
                x1_1, y1_1, x2_1, y2_1, id_1 = bbox1
                cx_1 = (x1_1 + x2_1) // 2
                cy_1 = (y1_1 + y2_1) // 2

                # People ID display
                cv2.rectangle(frame, (x1_1, y1_1), (x2_1, y2_1), (0, 255, 0), 2)
                cv2.circle(frame, (cx_1, cy_1), 4, (0, 0, 255), -1)
                cv2.putText(frame, str(id_1), (cx_1, cy_1), cv2.FONT_HERSHEY_COMPLEX, 0.8, (0, 255, 255), 2)
                
                for j, bbox2 in enumerate(bbox_id):
                    if i != j:  # Avoid self-comparison
                        x1_2, y1_2, x2_2, y2_2, id_2 = bbox2

                        distance = ((x1_1 - x1_2)**2 + (y1_1 - y1_2)**2)**0.5

                        # Calculate intersection area
                        x_overlap = max(0, min(x2_1, x2_2) - max(x1_1, x1_2)) #X-Axis Overlap
                        y_overlap = max(0, min(y2_1, y2_2) - max(y1_1, y1_2)) #Y-Axis Overlap
                        intersection_area = x_overlap * y_overlap
                        
                        # Calculate bounding box areas
                        area_bbox1 = (x2_1 - x1_1) * (y2_1 - y1_1)
                        area_bbox2 = (x2_2 - x1_2) * (y2_2 - y1_2)
                        
                        # Calculate overlap ratio
                        overlap_ratio1 = intersection_area / area_bbox1
                        overlap_ratio2 = intersection_area / area_bbox2
                        
                        # Raise warning if either overlap ratio exceeds threshold
                        if overlap_threshold_limit > overlap_ratio1 > overlap_threshold or overlap_threshold_limit > overlap_ratio2 > overlap_threshold:
                            cv2.putText(frame, "Alert: Collision!", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                            alert_shown = True
                            log_data.append({'Frame': count, 'Status': 'Collision', 'Timestamp': datetime.now(), 'Distance': distance})

                        elif distance < min_distance_threshold:
                            cv2.putText(frame, "Warning: Close proximity!", (50, 100), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 255), 2)
                            warning_shown = True
                            log_data.append({'Frame': count, 'Status': 'Close Proximity', 'Timestamp': datetime.now(), 'Distance': distance})
                            
                            
            if not warning_shown and not alert_shown:
                cv2.putText(frame, "Safe", (50, 150), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                log_data.append({'Frame': count, 'Status': 'Safe', 'Timestamp': datetime.now(), 'Distance': 'None'})
                
            out.write(frame)
            if cv2.waitKey(1) & 0xFF == 27:
                break

        # Release resources
        cap.release()
        out.release()
        cv2.destroyAllWindows()

        # Create DataFrame from log data
        log_df = pd.DataFrame(log_data)

        # Write DataFrame to Excel with formatting
        with pd.ExcelWriter('./static/output/log.xlsx', engine='openpyxl') as writer:
            log_df.to_excel(writer, index=False, startrow=12)

            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            for col in worksheet.iter_cols(min_row=1, max_row=8):
                for cell in col:
                    cell.fill = PatternFill(start_color='00425A', end_color='00425A', fill_type='solid')
            
            for col in worksheet.iter_cols(min_row=9, max_row=11):
                for cell in col:
                    cell.fill = PatternFill(start_color='1F8A70', end_color='1F8A70', fill_type='solid')

            for col in worksheet.iter_cols(min_row=12, max_row=12):
                for cell in col:
                    cell.fill = PatternFill(start_color='00425A', end_color='00425A', fill_type='solid')

            for row in worksheet.iter_rows(min_row=13, min_col=13, max_col=13, max_row=worksheet.max_row):
                cell = row[0] 
                if cell.value == 'Safe':
                    cell.font = safe_font
                elif cell.value == 'Collision':
                    cell.font = collision_font
                elif cell.value == 'Close Proximity':
                    cell.font = close_proximity_font

            # Add image at cell B2
            img = Image('./static/img/white logo.png')  # replace with your image path
            img.width = 100
            img.height = 100
            worksheet.add_image(img, 'B2')

            # Merge cells A10:D10 and write "Player Collision Prediction" in white color
            worksheet.merge_cells('A10:D10')
            cell = worksheet['A10']
            cell.value = "Player Collision Prediction"
            cell.font = Font(color="FFFFFF")  # Set font color to white

            # Align text to the center
            cell.alignment = Alignment(horizontal='center', vertical='center')


        print(log_df)


        # Load the excel file
        df = pd.read_excel('./static/output/log.xlsx', header=None, names=['Status'], skiprows=13, usecols="B")

        # Count the occurrences of each status
        status_counts = df['Status'].value_counts()

        # Write the counts to a new sheet in the same excel file
        with pd.ExcelWriter('./static/output/log.xlsx', engine='openpyxl', mode='a') as writer:
            status_counts.to_excel(writer, sheet_name='Graph')

        # Generate the pie chart
        plt.figure(figsize=(6,6))
        plt.pie(status_counts, labels=status_counts.index, autopct='%1.1f%%')
        plt.title('Pie Chart')
        plt.savefig('./static/output/pie_chart.png')

        # Load the workbook and select the sheet
        wb = openpyxl.load_workbook('./static/output/log.xlsx')
        ws = wb['Graph']

        # Add the pie chart image
        img_pie = openpyxl.drawing.image.Image('./static/output/pie_chart.png')
        img_pie.anchor = 'D2'  # Adjust the cell location as needed
        ws.add_image(img_pie)

        # Save the workbook
        wb.save('./static/output/log.xlsx')
        
        output_filename = './static/output'
        shutil.make_archive(output_filename, 'zip', './static/output')

        socketio.emit('status_update', {'status': 'success'})

    # with open('./static/temp/file.txt', 'w') as file:
    #     file.write('success')

    return send_file(f'{output_filename}.zip', as_attachment=True)


if __name__ == '__main__':
    socketio.run(app, host="0.0.0.0", port=8080, debug=True)
